import os
import aiohttp
import uuid
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from aiogram.types import FSInputFile, Message


# Функция для получения ответов формы через API
async def get_form_responses(form_id: str) -> dict:
    """Получает ответы формы через API"""
    url = f"https://test.upmini.app/api/forms/{form_id}/responses"
    
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(url) as response:
                if response.status == 200:
                    return await response.json()
                else:
                    raise Exception(f"API вернул статус {response.status}")
    except Exception as e:
        raise Exception(f"Ошибка при запросе к API: {str(e)}")


# Функция для генерации Excel файла с ответами
def generate_excel_file(responses_data: dict, form_id: str) -> str:
    """Генерирует Excel файл с ответами формы"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Responses_{form_id[:8]}"
    
    # Собираем все уникальные ключи из всех ответов
    all_keys = set()
    for response in responses_data.get('responses', []):
        if 'response_data' in response:
            all_keys.update(response['response_data'].keys())
    
    all_keys = sorted(list(all_keys))
    
    # Заголовки
    headers = ['ID', 'Form ID', 'Created At'] + all_keys
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Данные
    for row, response in enumerate(responses_data.get('responses', []), 2):
        ws.cell(row=row, column=1, value=response.get('id', ''))
        ws.cell(row=row, column=2, value=response.get('form_id', ''))
        ws.cell(row=row, column=3, value=response.get('created_at', ''))
        
        # Заполняем данные ответов
        response_data = response.get('response_data', {})
        for col, key in enumerate(all_keys, 4):
            value = response_data.get(key, '')  # Пустая ячейка если ключа нет
            ws.cell(row=row, column=col, value=value)
    
    # Автоподбор ширины колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем файл
    filename = f"form_responses_{form_id[:8]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = f"/tmp/{filename}"
    wb.save(filepath)
    return filepath


# Основная функция для обработки команды /form_res
async def handle_form_res_command(message: Message):
    """Обработчик команды /form_res для получения ответов формы"""
    try:
        # Извлекаем form_id из команды
        command_parts = message.text.split()
        if len(command_parts) != 2:
            await message.answer(
                "❌ Неверный формат команды. Используйте: /form_res <form_id>\n"
                "Пример: /form_res 0f913c81-cc46-42cd-9a0b-8458a8da2a2a"
            )
            return
        
        form_id = command_parts[1]
        
        # Проверяем, что form_id выглядит как UUID
        try:
            uuid.UUID(form_id)
        except ValueError:
            await message.answer("❌ Неверный формат UUID. Проверьте правильность form_id.")
            return
        
        # Отправляем сообщение о начале обработки
        processing_msg = await message.answer("⏳ Получаю ответы формы...")
        
        try:
            # Получаем данные с API
            responses_data = await get_form_responses(form_id)
            
            if not responses_data.get('responses'):
                await processing_msg.edit_text("❌ Ответы для данной формы не найдены.")
                return
            
            # Обновляем сообщение
            await processing_msg.edit_text("📊 Генерирую Excel файл...")
            
            # Генерируем Excel файл
            excel_filepath = generate_excel_file(responses_data, form_id)
            
            # Отправляем файл пользователю
            excel_file = FSInputFile(excel_filepath)
            await message.answer_document(
                document=excel_file,
                caption=f"📋 Ответы формы {form_id[:8]}...\n"
                       f"Всего ответов: {responses_data.get('total', 0)}"
            )
            
            # Удаляем временный файл
            try:
                os.remove(excel_filepath)
            except:
                pass
                
            await processing_msg.delete()
            
        except Exception as e:
            await processing_msg.edit_text(f"❌ Ошибка при получении данных: {str(e)}")
            
    except Exception as e:
        await message.answer(f"❌ Произошла ошибка: {str(e)}")
